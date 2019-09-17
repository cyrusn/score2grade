import yaml
import argparse
from openpyxl import load_workbook
from grading import Grading
from scores import Scores
from helper import isHeaderRow
from typing import Union

parser = argparse.ArgumentParser(
    description="Tool for convert exam score to DSE predicted grade",
    formatter_class=argparse.ArgumentDefaultsHelpFormatter,
)

parser.add_argument(
    "config", nargs="?", metavar="CONFIG", help="config file", default="config.yaml"
)

parser.add_argument("-s", "--score", nargs="?", help="score file path")
parser.add_argument("-u", "--student", nargs="?", help="students file path")
parser.add_argument("-g", "--grading", nargs="?", help="grading file path")
parser.add_argument("-o", "--output", nargs="?", help="output file path")

args = parser.parse_args()

config = yaml.safe_load(open(args.config, "r", encoding="utf-8"))

GRADING_FILE = args.grading if args.grading is not None else config["file"]["grading"]
STUDENTS_FILE = args.student if args.student is not None else config["file"]["students"]
SCORES_FILE = args.score if args.score is not None else config["file"]["scores"]
OUTPUT_FILE = args.output if args.output is not None else config["file"]["output"]


if __name__ == "__main__":
    grading_wb = load_workbook(GRADING_FILE)
    grading_ws = grading_wb.active
    grading = Grading(grading_ws)

    score_wb = load_workbook(SCORES_FILE)
    scores = Scores(score_wb, grading)

    student_wb = load_workbook(STUDENTS_FILE)
    student_sheet = student_wb.active

    classcode_column = student_sheet["A"]

    ELECTIVES = ["m2", "x2", "x1"]
    for i, elective in enumerate(ELECTIVES):
        col_index = 7 - i
        # insert a new column to the right of column with electives name
        # start from the far right column which is m2
        student_sheet.insert_cols(col_index)
        for classcode_cell in classcode_column:
            row_index = classcode_cell.row

            # subject code locates at the left column of the inserted column
            subject = student_sheet.cell(row=row_index, column=col_index - 1).value
            content: Union[int, str, None] = None

            if isHeaderRow(classcode_cell):
                content = f"{elective}_grade"
            elif subject is None:
                # mean student dropped the elective
                pass
            else:
                content = scores.get_grade(classcode_cell.value, subject)

            student_sheet.cell(row=row_index, column=col_index).value = content

    CORE_SUBJECTS = ["chi", "eng", "math", "ls"]
    for i, subject in enumerate(CORE_SUBJECTS):
        col_index = 4 + i
        student_sheet.insert_cols(col_index, amount=1)
        content = None
        for classcode_cell in classcode_column:
            classcode = classcode_cell.value
            row_index = classcode_cell.row
            if isHeaderRow(classcode_cell):
                content = subject
            elif subject is None:
                pass
            else:
                content = scores.get_grade(classcode, subject)

            student_sheet.cell(row=row_index, column=col_index).value = content

    student_wb.save(OUTPUT_FILE)
