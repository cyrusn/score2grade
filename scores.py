from grading import Grading
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from statistics import mean
from typing import List


class Scores:
    def __init__(
        self,
        score_wb: Workbook,
        grading: Grading,
        score_column_index=4,
        classcode_column_index=1,
    ):
        self.wb = score_wb
        self.grading = grading
        self.score_column_index = score_column_index
        self.classcode_column_letter = get_column_letter(classcode_column_index)

    def _get_grade(self, classcode: str, subject: str) -> Optional[int]:
        score_sheet = self.wb[subject]
        classcode_column = score_sheet[self.classcode_column_letter]
        for classcode_cell in classcode_column:
            if classcode_cell.value == classcode:
                score = score_sheet.cell(
                    row=classcode_cell.row, column=self.score_column_index
                ).value
                return self.grading.get_grade_by_score(subject, score)

        return None

    def _compute_cscb_grade(self, classcode: str) -> int:
        grades: List[int] = []
        for part in ["csc", "csb"]:
            grade = self._get_grade(classcode, part)
            if grade is not None:
                grades.append(grade)
        return int(round(mean(grades)))

    def get_grade(self, classcode: str, subject: str) -> Optional[int]:
        return (
            self._get_grade(classcode, subject)
            if subject != "cscb"
            else self._compute_cscb_grade(classcode)
        )

