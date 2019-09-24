from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import Optional, Union
from helper import bail


class Grading:
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.header_row = ws[1]
        self.subject_column = {
            cell.value: self.ws[get_column_letter(cell.column)]
            for cell in self.header_row
        }

    def get_grade_by_score(
        self, subject: str, score: Union[int, float]
    ) -> Optional[int]:
        # why cast score and grading range to `int`?
        score = int(round(score))

        if subject not in self.subject_column:
            bail(f"Subject ({subject}) not find")

        subject_column = self.subject_column[subject]
        for cell in subject_column[1:]:
            lower, upper = [int(n) for n in cell.value.split(", ")]

            isWithinRange = lower <= score and score <= upper
            if isWithinRange:
                return self.ws.cell(row=cell.row, column=1).value

        bail(f"Score[{score}] of the subject[{subject}] is out of range")
        return None


# # TODO: try update tests
# if __name__ == "__main__":
#     from openpyxl import load_workbook
#     import unittest

#     wb = load_workbook("./data/grading_ranges.xlsx")
#     ws = wb.active
#     g = Grading(ws)

#     class TestGrading(unittest.TestCase):
#         def test_equal(self):
#             test_data = (("bio", 84, 2), ("bafs", 129.1, 3), ("chem", 154, 5))
#             for data in test_data:
#                 result = g.get_grade_by_score(data[0], data[1])
#                 expected = data[2]
#                 self.assertEqual(result, expected)

#         def test_not_found(self):
#             test_data = (("abcd", 84, None), ("hijk", 129.1, None))
#             for data in test_data:
#                 result = g.get_grade_by_score(data[0], data[1])
#                 expected = data[2]
#                 self.assertEqual(result, expected)

#         def test_out_of_range(self):
#             with self.assertRaises(SystemExit) as cm:
#                 g.get_grade_by_score("bio", 221)
#             self.assertEqual(cm.exception.code, 1)

#     unittest.main()
